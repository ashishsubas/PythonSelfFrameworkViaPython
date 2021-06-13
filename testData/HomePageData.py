import openpyxl


class HomePageData:
    test_Homepage_data = [
        {"firstname": "Ashish", "emailId": "ashishsubas7@gmail.com", "password": "ashish", "gender": "Male"},
        {"firstname": "Subas", "emailId": "subas7@gmail.com", "password": "subas", "gender": "Female"}]


    @staticmethod
    def getTestData(testcaseName):
            book = openpyxl.load_workbook("C:\\Users\\Nitin\\Downloads\\PythonDemo.xlsx")
            sheet = book.active
            Dict = {}
            for i in range(1, sheet.max_row + 1):  ## to get rows
                if sheet.cell(row=i, column=1).value == testcaseName:
                    for j in range(2, sheet.max_column + 1):  ## to get columns
                        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

            return [Dict]

